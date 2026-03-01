"""
AGENTE 2: Verificacion de Formato de Reporte de Incompatibilidades
- Descarga adjuntos Excel (.xlsx/.xls) o Word (.docx/.doc) del correo
- Revisa TODOS los adjuntos y elige el mejor resultado
- Verifica que el documento inicie con datos de empresa y fecha
- Retorna resultado de validacion de formato
"""
import base64
import os
import re
from datetime import datetime, date

from config import TEMP_DIR, COMPANY_NAME


def verificar_formato_reporte(service, message_id, adjuntos, fecha_inicio, fecha_fin,
                              drive_service=None, links_drive=None):
    """
    Descarga y verifica el formato de TODOS los adjuntos del reporte de incompatibilidades.
    Tambien verifica archivos compartidos via links de Google Drive/Sheets.
    Retorna el mejor resultado encontrado entre todos los adjuntos y links.

    Args:
        service: Gmail API service
        message_id: ID del mensaje de Gmail
        adjuntos: Lista de adjuntos encontrados
        fecha_inicio: date object (lunes de la semana)
        fecha_fin: date object (domingo de la semana)
        drive_service: Google Drive API service (para descargar desde links)
        links_drive: Lista de links de Google Drive/Sheets encontrados en el correo

    Returns:
        dict con el mejor resultado de validacion encontrado
    """
    if not adjuntos and not links_drive:
        return {
            "formato_valido": False,
            "tipo_archivo": "sin_adjunto",
            "tiene_datos_empresa": False,
            "tiene_fecha": False,
            "fecha_documento": None,
            "fecha_correcta": False,
            "detalle": "Sin adjunto Excel/Word ni link de Drive",
        }

    os.makedirs(TEMP_DIR, exist_ok=True)

    # Verificar TODOS los adjuntos y elegir el mejor resultado
    resultados = []

    # 1. Verificar adjuntos directos del correo
    for adjunto in (adjuntos or []):
        filename = adjunto["filename"]
        attachment_id = adjunto["attachmentId"]

        try:
            resultado = _verificar_un_adjunto(
                service, message_id, attachment_id, filename, fecha_inicio, fecha_fin
            )
            if resultado:
                resultados.append(resultado)
        except Exception as e:
            resultados.append({
                "formato_valido": False,
                "tipo_archivo": "error",
                "tiene_datos_empresa": False,
                "tiene_fecha": False,
                "fecha_documento": None,
                "fecha_correcta": False,
                "detalle": f"Error en {filename}: {e}",
                "nombre_archivo": filename,
                "_score": 0,
            })

    # 2. Verificar archivos desde links de Google Drive/Sheets
    if drive_service and links_drive:
        for link in links_drive:
            try:
                resultado = _verificar_link_drive(
                    drive_service, link, fecha_inicio, fecha_fin
                )
                if resultado:
                    resultados.append(resultado)
            except Exception as e:
                resultados.append({
                    "formato_valido": False,
                    "tipo_archivo": "error",
                    "tiene_datos_empresa": False,
                    "tiene_fecha": False,
                    "fecha_documento": None,
                    "fecha_correcta": False,
                    "detalle": f"Error en link Drive: {e}",
                    "nombre_archivo": f"drive:{link.get('file_id', '?')}",
                    "_score": 0,
                })

    if not resultados:
        return {
            "formato_valido": False,
            "tipo_archivo": "error",
            "tiene_datos_empresa": False,
            "tiene_fecha": False,
            "fecha_documento": None,
            "fecha_correcta": False,
            "detalle": "No se pudo verificar ningun adjunto ni link",
        }

    # Elegir el mejor resultado (mayor score)
    mejor = max(resultados, key=lambda r: r.get("_score", 0))
    mejor.pop("_score", None)
    return mejor


def _verificar_un_adjunto(service, message_id, attachment_id, filename, fecha_inicio, fecha_fin):
    """Descarga y verifica un adjunto individual."""
    # Descargar adjunto
    try:
        att = (
            service.users()
            .messages()
            .attachments()
            .get(userId="me", messageId=message_id, id=attachment_id)
            .execute()
        )
        data = base64.urlsafe_b64decode(att["data"])
    except Exception as e:
        return {
            "formato_valido": False,
            "tipo_archivo": "error",
            "tiene_datos_empresa": False,
            "tiene_fecha": False,
            "fecha_documento": None,
            "fecha_correcta": False,
            "detalle": f"Error descargando {filename}: {e}",
            "nombre_archivo": filename,
            "_score": 0,
        }

    # Guardar temporalmente
    filepath = os.path.join(TEMP_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(data)

    try:
        # Extraer texto segun tipo
        if filename.lower().endswith((".xlsx", ".xls")):
            texto_inicio, fechas_celdas = _extraer_texto_excel(filepath)
            tipo_archivo = "excel"
        elif filename.lower().endswith((".docx", ".doc")):
            texto_inicio = _extraer_texto_word(filepath)
            fechas_celdas = []
            tipo_archivo = "word"
        else:
            return {
                "formato_valido": False,
                "tipo_archivo": "desconocido",
                "tiene_datos_empresa": False,
                "tiene_fecha": False,
                "fecha_documento": None,
                "fecha_correcta": False,
                "detalle": f"Tipo de archivo no soportado: {filename}",
                "nombre_archivo": filename,
                "_score": 0,
            }

        # Validar contenido
        resultado = _validar_contenido(
            texto_inicio, tipo_archivo, fecha_inicio, fecha_fin, filename, fechas_celdas
        )

        # Calcular score para elegir el mejor adjunto
        score = 0
        if resultado.get("formato_valido"):
            score += 10
        if resultado.get("tiene_datos_empresa"):
            score += 5
        if resultado.get("tiene_fecha"):
            score += 3
        if resultado.get("fecha_correcta"):
            score += 20
        resultado["_score"] = score

        return resultado

    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass


def _verificar_link_drive(drive_service, link, fecha_inicio, fecha_fin):
    """
    Descarga un archivo desde Google Drive/Sheets y verifica su formato.

    Args:
        drive_service: Google Drive API v3 service
        link: dict con file_id, url, tipo
        fecha_inicio: date
        fecha_fin: date

    Returns:
        dict con resultado de validacion
    """
    import io

    file_id = link["file_id"]
    tipo_link = link.get("tipo", "google_drive")

    # Obtener metadata del archivo
    try:
        file_meta = drive_service.files().get(
            fileId=file_id,
            fields="id, name, mimeType",
            supportsAllDrives=True,
        ).execute()
    except Exception as e:
        return {
            "formato_valido": False,
            "tipo_archivo": "error",
            "tiene_datos_empresa": False,
            "tiene_fecha": False,
            "fecha_documento": None,
            "fecha_correcta": False,
            "detalle": f"Error accediendo archivo Drive ({file_id}): {e}",
            "nombre_archivo": f"drive:{file_id}",
            "_score": 0,
        }

    filename = file_meta.get("name", f"drive_{file_id}")
    mime_type = file_meta.get("mimeType", "")

    # Descargar segun tipo de archivo
    # IMPORTANTE: Verificar primero los formatos nativos de Google (vnd.google-apps.*)
    # ya que estos requieren export. Los archivos binarios subidos a Drive usan get_media.
    try:
        if mime_type == "application/vnd.google-apps.spreadsheet":
            # Google Sheets NATIVO: exportar como xlsx
            response = drive_service.files().export(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ).execute()
            data = response
            filename_local = f"{filename}.xlsx"
            tipo_archivo = "excel"
        elif mime_type == "application/vnd.google-apps.document":
            # Google Docs NATIVO: exportar como docx
            response = drive_service.files().export(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ).execute()
            data = response
            filename_local = f"{filename}.docx"
            tipo_archivo = "word"
        elif mime_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "application/vnd.ms-excel.sheet.macroEnabled.12",
        ]:
            # Excel binario subido a Drive: descargar directo
            response = drive_service.files().get_media(fileId=file_id).execute()
            data = response
            filename_local = filename if filename.lower().endswith((".xlsx", ".xls")) else f"{filename}.xlsx"
            tipo_archivo = "excel"
        elif mime_type in [
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ]:
            # Word binario subido a Drive: descargar directo
            response = drive_service.files().get_media(fileId=file_id).execute()
            data = response
            filename_local = filename
            tipo_archivo = "word"
        else:
            return {
                "formato_valido": False,
                "tipo_archivo": "desconocido",
                "tiene_datos_empresa": False,
                "tiene_fecha": False,
                "fecha_documento": None,
                "fecha_correcta": False,
                "detalle": f"Tipo no soportado: {mime_type} ({filename})",
                "nombre_archivo": filename,
                "_score": 0,
            }
    except Exception as e:
        return {
            "formato_valido": False,
            "tipo_archivo": "error",
            "tiene_datos_empresa": False,
            "tiene_fecha": False,
            "fecha_documento": None,
            "fecha_correcta": False,
            "detalle": f"Error descargando de Drive: {e}",
            "nombre_archivo": filename,
            "_score": 0,
        }

    # Guardar temporalmente
    filepath = os.path.join(TEMP_DIR, filename_local)
    with open(filepath, "wb") as f:
        f.write(data)

    try:
        if tipo_archivo == "excel":
            texto_inicio, fechas_celdas = _extraer_texto_excel(filepath)
        elif tipo_archivo == "word":
            texto_inicio = _extraer_texto_word(filepath)
            fechas_celdas = []
        else:
            return None

        resultado = _validar_contenido(
            texto_inicio, tipo_archivo, fecha_inicio, fecha_fin, filename, fechas_celdas
        )

        # Calcular score
        score = 0
        if resultado.get("formato_valido"):
            score += 10
        if resultado.get("tiene_datos_empresa"):
            score += 5
        if resultado.get("tiene_fecha"):
            score += 3
        if resultado.get("fecha_correcta"):
            score += 20
        resultado["_score"] = score

        # Marcar que viene de link de Drive
        resultado["nombre_archivo"] = f"{filename} (link Drive)"

        return resultado

    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass


def _extraer_texto_excel(filepath):
    """
    Extrae las primeras celdas de texto de un archivo Excel.
    Tambien extrae fechas nativas (datetime objects) de las celdas.

    Returns:
        tuple (texto_str, fechas_celdas_list)
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return "", []

        textos = []
        fechas_celdas = []

        # Leer las primeras 30 filas, hasta 15 columnas
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=15, values_only=True):
            for cell in row:
                if cell is not None:
                    # Detectar fechas nativas de Excel (datetime objects)
                    if isinstance(cell, datetime):
                        fechas_celdas.append(cell.date())
                        textos.append(cell.strftime("%d/%m/%Y"))
                    elif isinstance(cell, date):
                        fechas_celdas.append(cell)
                        textos.append(cell.strftime("%d/%m/%Y"))
                    else:
                        textos.append(str(cell))
            if len(" ".join(textos)) > 2000:
                break

        wb.close()
        return " ".join(textos)[:2000], fechas_celdas
    except ImportError:
        return "[ERROR: Instalar openpyxl para leer archivos Excel]", []
    except Exception as e:
        return f"[ERROR: {e}]", []


def _extraer_texto_word(filepath):
    """Extrae las primeras lineas de texto de un documento Word."""
    try:
        from docx import Document

        doc = Document(filepath)
        lineas = []
        for para in doc.paragraphs[:20]:  # Primeros 20 parrafos
            texto = para.text.strip()
            if texto:
                lineas.append(texto)
            if len("\n".join(lineas)) > 1000:
                break
        return "\n".join(lineas)[:1000]
    except ImportError:
        return "[ERROR: Instalar python-docx para leer archivos Word]"
    except Exception as e:
        return f"[ERROR: {e}]"


def _validar_contenido(texto_inicio, tipo_archivo, fecha_inicio, fecha_fin, filename, fechas_celdas=None):
    """Valida que el texto del documento contenga datos de empresa y fecha."""
    texto_lower = texto_inicio.lower()

    # Verificar datos de empresa (keywords desde config)
    from config import KEYWORDS_EMPRESA
    tiene_datos_empresa = any(kw in texto_lower for kw in KEYWORDS_EMPRESA)

    # --- EXTRACCION DE FECHA ---
    fecha_documento = None
    fecha_correcta = False
    tiene_alguna_fecha = False

    # Prioridad 1: Fechas nativas de Excel (datetime objects - mas confiables)
    if fechas_celdas:
        tiene_alguna_fecha = True
        for fecha_celda in fechas_celdas:
            if fecha_inicio <= fecha_celda <= fecha_fin:
                fecha_documento = fecha_celda.strftime("%d/%m/%Y")
                fecha_correcta = True
                break
        # Si ninguna fecha cae en la semana, usar la primera
        if not fecha_documento and fechas_celdas:
            fecha_documento = fechas_celdas[0].strftime("%d/%m/%Y")

    # Prioridad 2: Fechas en el nombre del archivo (muy confiable)
    if not fecha_correcta:
        fecha_filename = _extraer_fecha_filename(filename, fecha_inicio, fecha_fin)
        if fecha_filename:
            tiene_alguna_fecha = True
            if not fecha_documento or not fecha_correcta:
                fecha_documento = fecha_filename["fecha_str"]
                fecha_correcta = fecha_filename["correcta"]

    # Prioridad 3: Fechas en texto con formato dd/mm/yyyy o similar
    if not tiene_alguna_fecha or not fecha_correcta:
        patron_fecha = r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})'
        fechas_encontradas = re.findall(patron_fecha, texto_inicio)

        if fechas_encontradas:
            tiene_alguna_fecha = True

            # Buscar la primera fecha que sea una fecha real valida
            for dia, mes, anio in fechas_encontradas:
                try:
                    if len(anio) == 2:
                        anio_full = f"20{anio}"
                    else:
                        anio_full = anio

                    d = int(dia)
                    m = int(mes)
                    a = int(anio_full)

                    # Validar que sea una fecha real (no un numero random)
                    if not (1 <= d <= 31 and 1 <= m <= 12 and 2020 <= a <= 2030):
                        continue

                    fecha_parsed = date(a, m, d)
                    fecha_str = f"{dia.zfill(2)}/{mes.zfill(2)}/{anio_full}"

                    if fecha_inicio <= fecha_parsed <= fecha_fin:
                        fecha_documento = fecha_str
                        fecha_correcta = True
                        break
                    elif not fecha_documento:
                        # Guardar la primera fecha valida como referencia
                        fecha_documento = fecha_str
                except (ValueError, OverflowError):
                    continue

    # Determinar formato valido
    formato_valido = tiene_datos_empresa and tiene_alguna_fecha

    # Construir detalle
    rango_str = f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    detalles = []
    if not tiene_datos_empresa:
        detalles.append("Sin datos de empresa al inicio")
    if not tiene_alguna_fecha:
        detalles.append("Sin fecha en el documento")
    elif not fecha_correcta:
        detalles.append(f"Fecha documento: {fecha_documento} (esperado: semana {rango_str})")

    detalle = "; ".join(detalles) if detalles else "Formato correcto"

    return {
        "formato_valido": formato_valido,
        "tipo_archivo": tipo_archivo,
        "tiene_datos_empresa": tiene_datos_empresa,
        "tiene_fecha": tiene_alguna_fecha,
        "fecha_documento": fecha_documento,
        "fecha_correcta": fecha_correcta,
        "detalle": detalle,
        "nombre_archivo": filename,
    }


def _extraer_fecha_filename(filename, fecha_inicio, fecha_fin):
    """
    Intenta extraer una fecha del nombre del archivo.
    Soporta formatos como:
      - 18-02-2026, 18_02_26, 16_02_26
      - (18-02-2026)
    """
    # Patron: dd-mm-yyyy, dd_mm_yyyy, dd-mm-yy, dd_mm_yy
    patron = r'(\d{1,2})[\-_](\d{1,2})[\-_](\d{2,4})'
    matches = re.findall(patron, filename)

    for dia, mes, anio in matches:
        try:
            if len(anio) == 2:
                anio_full = f"20{anio}"
            else:
                anio_full = anio

            d = int(dia)
            m = int(mes)
            a = int(anio_full)

            if not (1 <= d <= 31 and 1 <= m <= 12 and 2020 <= a <= 2030):
                continue

            fecha_parsed = date(a, m, d)
            fecha_str = f"{dia.zfill(2)}/{mes.zfill(2)}/{anio_full}"
            correcta = fecha_inicio <= fecha_parsed <= fecha_fin

            return {"fecha_str": fecha_str, "correcta": correcta, "date": fecha_parsed}
        except (ValueError, OverflowError):
            continue

    return None
